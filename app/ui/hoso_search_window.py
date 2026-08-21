"""Cửa sổ tìm và tải hồ sơ trong bộ công cụ PDF."""

from __future__ import annotations

import copy
import ctypes
import os
import shutil
import time
from ctypes import wintypes
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QObject, QSignalBlocker, QThread, Qt, Signal, Slot
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)


MAX_FOLDERS = 10
ERROR_MORE_DATA = 234
MAX_PREFERRED_LENGTH = 0xFFFFFFFF
STYPE_DISKTREE = 0


class SHARE_INFO_1(ctypes.Structure):
    _fields_ = [
        ("shi1_netname", wintypes.LPWSTR),
        ("shi1_type", wintypes.DWORD),
        ("shi1_remark", wintypes.LPWSTR),
    ]


class SearchWorker(QObject):
    """Quét tên tệp ở thread nền; tuyệt đối không cập nhật giao diện tại đây."""

    current_folder = Signal(str)
    scanned = Signal(int, int)
    progress = Signal(int)
    matched = Signal(dict)
    log = Signal(str)
    finished = Signal(bool, int, int)

    def __init__(self, folders: list[str], records: list[dict]):
        super().__init__()
        self.folders = folders
        self.records = records
        self._stop_requested = False

    def request_stop(self):
        self._stop_requested = True

    @Slot()
    def run(self):
        total_files = 0
        found_files = 0
        folder_count = len(self.folders)

        for folder_number, folder in enumerate(self.folders, start=1):
            if self._stop_requested:
                break

            self.current_folder.emit(folder)
            self.log.emit(f"Bắt đầu quét {folder}")
            self.progress.emit(round((folder_number - 1) / folder_count * 100))

            def report_walk_error(error):
                self.log.emit(f"Không thể đọc: {error}")

            for root, _directories, files in os.walk(folder, onerror=report_walk_error):
                if self._stop_requested:
                    break

                for filename in files:
                    if self._stop_requested:
                        break

                    total_files += 1
                    filename_key = filename.casefold()
                    full_path = os.path.join(root, filename)

                    for record in self.records:
                        if record["found"]:
                            continue

                        mode = record["mode"]
                        matches = (
                            mode == "both"
                            and record["so_key"] in filename_key
                            and record["ten_key"] in filename_key
                        ) or (
                            mode == "so" and record["so_key"] in filename_key
                        ) or (
                            mode == "ten" and record["ten_key"] in filename_key
                        )

                        if not matches:
                            continue

                        record["found"] = True
                        record["file_name"] = filename
                        record["file_path"] = full_path
                        found_files += 1
                        self.matched.emit(copy.deepcopy(record))
                        self.log.emit(f"Tìm thấy: {filename}")
                        # Một tệp chỉ được gán cho một hồ sơ.
                        break

                    if total_files % 200 == 0:
                        self.scanned.emit(total_files, found_files)

            self.progress.emit(round(folder_number / folder_count * 100))
            self.log.emit(f"Hoàn thành {folder}")

        self.scanned.emit(total_files, found_files)
        self.finished.emit(self._stop_requested, total_files, found_files)


class HoSoSearchWindow(QMainWindow):
    """Phiên bản PySide6 của HoSoSearch, chạy trong cùng ứng dụng PDF."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        self.setWindowTitle("HoSo PDF Toolkit — Tìm hồ sơ")
        self.resize(1420, 900)
        self.setMinimumSize(1120, 720)

        self.records: list[dict] = []
        self.total_files = 0
        self.found_files = 0
        self.started_at: float | None = None
        self._thread: QThread | None = None
        self._worker: SearchWorker | None = None
        self._close_requested = False
        self._build_ui()

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        layout.setContentsMargins(10, 10, 10, 10)

        connection_group = QGroupBox("Nguồn dữ liệu")
        connection_layout = QGridLayout(connection_group)
        connection_layout.setColumnStretch(1, 1)

        self.machine_edit = QLineEdit("AK22-20220125YJ")
        self.share_combo = QComboBox()
        self.share_combo.currentIndexChanged.connect(self._change_share)
        self.excel_edit = QLineEdit()
        self.save_edit = QLineEdit()

        self.connect_button = QPushButton("Kết nối")
        self.connect_button.clicked.connect(self.connect_to_machine)
        choose_excel = QPushButton("Chọn Excel")
        choose_excel.clicked.connect(self.choose_excel)
        choose_save = QPushButton("Chọn nơi lưu")
        choose_save.clicked.connect(self.choose_save_folder)

        connection_layout.addWidget(QLabel("Tên máy/IP"), 0, 0)
        connection_layout.addWidget(self.machine_edit, 0, 1)
        connection_layout.addWidget(QLabel("Chia sẻ"), 0, 2)
        connection_layout.addWidget(self.share_combo, 0, 3)
        connection_layout.addWidget(self.connect_button, 0, 4)
        connection_layout.addWidget(QLabel("Excel"), 1, 0)
        connection_layout.addWidget(self.excel_edit, 1, 1, 1, 3)
        connection_layout.addWidget(choose_excel, 1, 4)
        connection_layout.addWidget(QLabel("Nơi lưu"), 2, 0)
        connection_layout.addWidget(self.save_edit, 2, 1, 1, 3)
        connection_layout.addWidget(choose_save, 2, 4)
        layout.addWidget(connection_group)

        folders_group = QGroupBox("Thư mục trên máy đã kết nối")
        folders_layout = QVBoxLayout(folders_group)
        self.folder_list = QListWidget()
        self.folder_list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.folder_list.itemSelectionChanged.connect(self.limit_folder_selection)
        folders_layout.addWidget(self.folder_list)
        self.folder_count_label = QLabel(f"Đã chọn: 0 / {MAX_FOLDERS}")
        folders_layout.addWidget(self.folder_count_label)
        layout.addWidget(folders_group)

        actions = QHBoxLayout()
        self.read_button = QPushButton("Đọc Excel")
        self.search_button = QPushButton("Bắt đầu tìm")
        self.stop_button = QPushButton("Dừng")
        self.download_button = QPushButton("Tải hồ sơ")
        self.export_button = QPushButton("Xuất Excel")
        self.stop_button.setEnabled(False)
        self.read_button.clicked.connect(self.read_excel)
        self.search_button.clicked.connect(self.start_search)
        self.stop_button.clicked.connect(self.stop_search)
        self.download_button.clicked.connect(self.download_found_files)
        self.export_button.clicked.connect(self.export_excel)
        for button in (
            self.read_button,
            self.search_button,
            self.stop_button,
            self.download_button,
            self.export_button,
        ):
            actions.addWidget(button)
        actions.addStretch()
        layout.addLayout(actions)

        progress_group = QGroupBox("Tiến trình")
        progress_layout = QGridLayout(progress_group)
        self.progress_bar = QProgressBar()
        self.current_folder_label = QLabel("Đang quét: -")
        self.total_label = QLabel("Đã quét: 0 tệp")
        self.found_label = QLabel("Đã tìm thấy: 0")
        self.speed_label = QLabel("Tốc độ: 0 tệp/s")
        self.eta_label = QLabel("ETA: --")
        progress_layout.addWidget(self.progress_bar, 0, 0, 1, 3)
        progress_layout.addWidget(self.current_folder_label, 1, 0, 1, 3)
        progress_layout.addWidget(self.total_label, 2, 0)
        progress_layout.addWidget(self.found_label, 2, 1)
        progress_layout.addWidget(self.speed_label, 2, 2)
        progress_layout.addWidget(self.eta_label, 3, 0, 1, 3)
        layout.addWidget(progress_group)

        self.results_table = QTableWidget(0, 6)
        self.results_table.setHorizontalHeaderLabels(
            ["STT", "Số phát hành", "Tên chủ", "Tên tệp", "Đường dẫn", "Trạng thái"]
        )
        self.results_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.results_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.itemDoubleClicked.connect(self.open_result_file)
        header = self.results_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.results_table, 1)

        self.log = QPlainTextEdit()
        self.log.setReadOnly(True)
        self.log.setMaximumBlockCount(1000)
        self.log.setFixedHeight(130)
        layout.addWidget(QLabel("Nhật ký"))
        layout.addWidget(self.log)

    @staticmethod
    def get_remote_shares(machine: str) -> list[str]:
        if os.name != "nt":
            raise OSError("Tính năng đọc thư mục chia sẻ chỉ hỗ trợ Windows.")

        netapi32 = ctypes.WinDLL("Netapi32.dll")
        enum_shares = netapi32.NetShareEnum
        enum_shares.argtypes = [
            wintypes.LPCWSTR,
            wintypes.DWORD,
            ctypes.POINTER(ctypes.c_void_p),
            wintypes.DWORD,
            ctypes.POINTER(wintypes.DWORD),
            ctypes.POINTER(wintypes.DWORD),
            ctypes.POINTER(wintypes.DWORD),
        ]
        enum_shares.restype = wintypes.DWORD
        free_buffer = netapi32.NetApiBufferFree
        free_buffer.argtypes = [ctypes.c_void_p]
        free_buffer.restype = wintypes.DWORD

        shares: list[str] = []
        resume_handle = wintypes.DWORD(0)
        while True:
            buffer = ctypes.c_void_p()
            entries_read = wintypes.DWORD()
            total_entries = wintypes.DWORD()
            status = enum_shares(
                rf"\\{machine}",
                1,
                ctypes.byref(buffer),
                MAX_PREFERRED_LENGTH,
                ctypes.byref(entries_read),
                ctypes.byref(total_entries),
                ctypes.byref(resume_handle),
            )
            try:
                if status not in (0, ERROR_MORE_DATA):
                    raise OSError(status, f"Không thể đọc thư mục chia sẻ của máy {machine}.")
                if buffer and entries_read.value:
                    info = ctypes.cast(buffer, ctypes.POINTER(SHARE_INFO_1))
                    for index in range(entries_read.value):
                        share = info[index]
                        share_type = share.shi1_type & 0xFF
                        share_name = share.shi1_netname
                        if (
                            share_type == STYPE_DISKTREE
                            and share_name
                            and share_name.casefold() not in {"admin$", "print$"}
                        ):
                            shares.append(share_name)
            finally:
                if buffer:
                    free_buffer(buffer)

            if status == 0:
                break
        return sorted(set(shares), key=str.casefold)

    def connect_to_machine(self):
        machine = self.machine_edit.text().strip().strip("\\/")
        if not machine:
            QMessageBox.warning(self, "Thiếu thông tin", "Hãy nhập tên máy hoặc IP.")
            return

        self.connect_button.setEnabled(False)
        self.share_combo.clear()
        self.folder_list.clear()
        try:
            shares = self.get_remote_shares(machine)
            available_shares: list[tuple[str, list[str]]] = []
            for share in shares:
                root = rf"\\{machine}\{share}"
                try:
                    folders = [
                        entry.path
                        for entry in sorted(os.scandir(root), key=lambda item: item.name.casefold())
                        if entry.is_dir()
                    ]
                except OSError as error:
                    self.add_log(f"Không thể mở {root}: {error}")
                    continue
                if folders:
                    available_shares.append((share, folders))
        except OSError as error:
            QMessageBox.critical(
                self,
                "Không thể kết nối",
                f"Không thể đọc thư mục chia sẻ của máy '{machine}'.\n\n{error}\n\n"
                "Hãy kiểm tra máy đang bật, cùng mạng và đã cấp quyền chia sẻ tệp.",
            )
            return
        finally:
            self.connect_button.setEnabled(True)

        if not available_shares:
            QMessageBox.warning(
                self,
                "Không có thư mục",
                f"Máy '{machine}' không có thư mục chia sẻ nào có thể truy cập.",
            )
            return

        for share, folders in available_shares:
            self.share_combo.addItem(share, folders)
        self._change_share()
        self.add_log(f"Đã kết nối {machine}; tìm thấy {len(available_shares)} thư mục chia sẻ.")

    def _change_share(self, _index: int | None = None):
        self.folder_list.clear()
        folders = self.share_combo.currentData() or []
        for path in folders:
            item = QListWidgetItem(os.path.basename(path))
            item.setData(Qt.ItemDataRole.UserRole, path)
            item.setToolTip(path)
            self.folder_list.addItem(item)
        self.folder_count_label.setText(f"Đã chọn: 0 / {MAX_FOLDERS}")

    def limit_folder_selection(self):
        selected = self.folder_list.selectedItems()
        if len(selected) > MAX_FOLDERS:
            current = self.folder_list.currentItem()
            rejected = current if current in selected else selected[-1]
            with QSignalBlocker(self.folder_list):
                rejected.setSelected(False)
            selected = self.folder_list.selectedItems()
            QMessageBox.warning(self, "Giới hạn thư mục", f"Chỉ được chọn tối đa {MAX_FOLDERS} thư mục.")
        self.folder_count_label.setText(f"Đã chọn: {len(selected)} / {MAX_FOLDERS}")

    def choose_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn tệp Excel", self.excel_edit.text(), "Excel (*.xlsx *.xls)"
        )
        if path:
            self.excel_edit.setText(path)

    def choose_save_folder(self):
        path = QFileDialog.getExistingDirectory(self, "Chọn nơi lưu", self.save_edit.text())
        if path:
            self.save_edit.setText(path)
            self.add_log(f"Nơi lưu: {path}")

    def read_excel(self):
        path = self.excel_edit.text().strip()
        if not path:
            QMessageBox.warning(self, "Thiếu tệp", "Hãy chọn tệp Excel trước.")
            return

        try:
            workbook = load_workbook(path, data_only=True)
            worksheet = workbook.active
            columns: dict[str, int] = {}
            for cell in worksheet[1]:
                if cell.value is None:
                    continue
                value = str(cell.value).strip().casefold()
                if value in {"stt", "số phát hành", "tên chủ"}:
                    columns[value] = cell.column

            if "số phát hành" not in columns and "tên chủ" not in columns:
                raise ValueError("Không tìm thấy cột 'Số phát hành' hoặc 'Tên chủ'.")

            records: list[dict] = []
            for row in range(2, worksheet.max_row + 1):
                def cell_text(column_name: str) -> str:
                    column = columns.get(column_name)
                    value = worksheet.cell(row=row, column=column).value if column else None
                    return "" if value is None else str(value).strip()

                stt = cell_text("stt")
                so = cell_text("số phát hành")
                ten = cell_text("tên chủ")
                if not so and not ten:
                    continue
                records.append(
                    {
                        "index": len(records),
                        "stt": stt,
                        "so": so,
                        "ten": ten,
                        "so_key": so.casefold(),
                        "ten_key": ten.casefold(),
                        "mode": "both" if so and ten else "so" if so else "ten",
                        "found": False,
                        "file_name": "",
                        "file_path": "",
                    }
                )
            self.records = records
            self._reset_results()
            self.add_log(f"Đã đọc {len(records)} hồ sơ từ Excel.")
            QMessageBox.information(self, "Hoàn thành", f"Đã đọc {len(records)} hồ sơ.")
        except Exception as error:
            QMessageBox.critical(self, "Không thể đọc Excel", str(error))

    def _reset_results(self):
        for record in self.records:
            record.update(found=False, file_name="", file_path="")
        self.results_table.setRowCount(0)
        self.total_files = 0
        self.found_files = 0
        self.progress_bar.setValue(0)
        self.current_folder_label.setText("Đang quét: -")
        self.total_label.setText("Đã quét: 0 tệp")
        self.found_label.setText("Đã tìm thấy: 0")
        self.speed_label.setText("Tốc độ: 0 tệp/s")
        self.eta_label.setText("ETA: --")

    def start_search(self):
        if self._thread and self._thread.isRunning():
            return
        if not self.records:
            QMessageBox.warning(self, "Chưa có dữ liệu", "Hãy đọc Excel trước.")
            return
        folders = [
            item.data(Qt.ItemDataRole.UserRole) for item in self.folder_list.selectedItems()
        ]
        if not folders:
            QMessageBox.warning(self, "Chưa chọn thư mục", "Hãy chọn ít nhất một thư mục để quét.")
            return

        self._reset_results()
        self.started_at = time.monotonic()
        self.search_button.setEnabled(False)
        self.read_button.setEnabled(False)
        self.stop_button.setEnabled(True)

        thread = QThread(self)
        worker = SearchWorker(folders, copy.deepcopy(self.records))
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.current_folder.connect(self._show_current_folder)
        worker.scanned.connect(self._show_status)
        worker.progress.connect(self.progress_bar.setValue)
        worker.matched.connect(self._add_match)
        worker.log.connect(self.add_log)
        worker.finished.connect(self._search_finished)
        worker.finished.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(self._clear_worker)
        self._thread = thread
        self._worker = worker
        thread.start()
        self.add_log(f"Bắt đầu quét {len(folders)} thư mục.")

    def stop_search(self):
        if self._worker:
            self._worker.request_stop()
            self.stop_button.setEnabled(False)
            self.add_log("Đã yêu cầu dừng quét.")

    def _show_current_folder(self, folder: str):
        self.current_folder_label.setText(f"Đang quét: {folder}")

    def _show_status(self, total: int, found: int):
        self.total_files = total
        self.found_files = found
        elapsed = max(0.001, time.monotonic() - (self.started_at or time.monotonic()))
        speed = total / elapsed
        self.total_label.setText(f"Đã quét: {total:,} tệp")
        self.found_label.setText(f"Đã tìm thấy: {found}")
        self.speed_label.setText(f"Tốc độ: {speed:,.0f} tệp/s")
        self.eta_label.setText("ETA: sẽ cập nhật theo từng thư mục")

    def _add_match(self, record: dict):
        self.records[record["index"]].update(record)
        row = self.results_table.rowCount()
        self.results_table.insertRow(row)
        for column, value in enumerate(
            (
                record["stt"],
                record["so"],
                record["ten"],
                record["file_name"],
                record["file_path"],
                "Đã tìm",
            )
        ):
            self.results_table.setItem(row, column, QTableWidgetItem(str(value)))

    def _search_finished(self, cancelled: bool, total: int, found: int):
        self._show_status(total, found)
        self.progress_bar.setValue(100)
        self.search_button.setEnabled(True)
        self.read_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        message = "Đã dừng quét." if cancelled else "Đã quét xong."
        self.add_log(f"{message} Tìm thấy {found} / {len(self.records)} hồ sơ.")

    def _clear_worker(self):
        thread = self._thread
        self._thread = None
        self._worker = None
        if thread:
            thread.deleteLater()
        if self._close_requested:
            self.close()

    def open_result_file(self, item: QTableWidgetItem):
        row = item.row()
        path_item = self.results_table.item(row, 4)
        if not path_item:
            return
        path = path_item.text()
        try:
            os.startfile(path)
        except OSError:
            os.system(f'explorer /select,"{path}"')

    def download_found_files(self):
        records = [record for record in self.records if record["found"]]
        if not records:
            QMessageBox.warning(self, "Không có hồ sơ", "Chưa có hồ sơ nào được tìm thấy.")
            return
        destination_folder = self.save_edit.text().strip()
        if not destination_folder:
            QMessageBox.warning(self, "Chưa có nơi lưu", "Hãy chọn nơi lưu trước.")
            return

        try:
            os.makedirs(destination_folder, exist_ok=True)
        except OSError as error:
            QMessageBox.critical(self, "Không thể tạo thư mục", str(error))
            return

        copied = failed = 0
        for record in records:
            source = record["file_path"]
            if not source or not os.path.isfile(source):
                failed += 1
                continue
            name = record["file_name"]
            destination = os.path.join(destination_folder, name)
            stem, extension = os.path.splitext(name)
            suffix = 1
            while os.path.exists(destination):
                destination = os.path.join(destination_folder, f"{stem} ({suffix}){extension}")
                suffix += 1
            try:
                shutil.copy2(source, destination)
                copied += 1
            except OSError as error:
                failed += 1
                self.add_log(f"Không thể tải {name}: {error}")

        self.add_log(f"Đã tải {copied} tệp; lỗi {failed} tệp.")
        QMessageBox.information(self, "Hoàn thành", f"Đã tải: {copied}\nLỗi: {failed}")

    def export_excel(self):
        if not self.records:
            QMessageBox.warning(self, "Không có dữ liệu", "Hãy đọc Excel trước.")
            return
        destination_folder = self.save_edit.text().strip()
        if not destination_folder:
            QMessageBox.warning(self, "Chưa có nơi lưu", "Hãy chọn nơi lưu trước.")
            return

        try:
            os.makedirs(destination_folder, exist_ok=True)
            output_path = os.path.join(
                destination_folder,
                datetime.now().strftime("KetQua_TimKiem_%Y%m%d_%H%M%S.xlsx"),
            )
            workbook = Workbook()
            found_sheet = workbook.active
            found_sheet.title = "DaTimThay"
            missing_sheet = workbook.create_sheet("KhongTimThay")
            summary_sheet = workbook.create_sheet("ThongKe")
            self._write_header(
                found_sheet,
                ["STT", "Số phát hành", "Tên chủ", "Tên tệp", "Đường dẫn", "Trạng thái"],
                "92D050",
            )
            self._write_header(
                missing_sheet,
                ["STT", "Số phát hành", "Tên chủ", "Trạng thái"],
                "FF9999",
            )
            found_count = 0
            for record in self.records:
                if record["found"]:
                    found_count += 1
                    found_sheet.append(
                        [
                            record["stt"],
                            record["so"],
                            record["ten"],
                            record["file_name"],
                            record["file_path"],
                            "Đã tìm",
                        ]
                    )
                else:
                    missing_sheet.append([record["stt"], record["so"], record["ten"], "Không tìm"])
            summary_rows = [
                ("Tổng số hồ sơ", len(self.records)),
                ("Đã tìm thấy", found_count),
                ("Không tìm thấy", len(self.records) - found_count),
                ("Tổng số tệp đã quét", self.total_files),
                ("Máy", self.machine_edit.text().strip()),
                ("Ngày xuất", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ]
            for label, value in summary_rows:
                summary_sheet.append([label, value])
            self._fit_columns(workbook)
            workbook.save(output_path)
        except Exception as error:
            QMessageBox.critical(self, "Không thể xuất Excel", str(error))
            return

        self.add_log(f"Đã xuất Excel: {output_path}")
        QMessageBox.information(self, "Hoàn thành", f"Đã xuất:\n{output_path}")
        try:
            os.startfile(output_path)
        except OSError:
            pass

    @staticmethod
    def _write_header(sheet, headers: list[str], color: str):
        for column, text in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column, value=text)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor=color)

    @staticmethod
    def _fit_columns(workbook: Workbook):
        for sheet in workbook.worksheets:
            for column_cells in sheet.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width + 5, 80)

    def add_log(self, message: str):
        self.log.appendPlainText(f"[{time.strftime('%H:%M:%S')}] {message}")

    def closeEvent(self, event):
        if self._thread and self._thread.isRunning():
            self._close_requested = True
            self.stop_search()
            self.add_log("Đang dừng quét trước khi đóng cửa sổ.")
            event.ignore()
            return
        event.accept()
